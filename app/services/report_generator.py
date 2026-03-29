"""Generador de reportes PDF y Excel."""
import csv
import io
from datetime import date
from decimal import Decimal
from typing import Optional, List, Dict
from sqlalchemy.orm import Session
from sqlalchemy import func, case

from app.models.balance import Balance
from app.models.disbursement import Disbursement
from app.models.contract import Contract
from app.models.creditor import Creditor
from app.models.payment import PaymentSchedule
from app.models.covenant import Covenant, CovenantTracking
from app.services.calculation_engine import CalculationEngine


class ReportGenerator:
    """Genera reportes del portafolio de deuda."""

    @staticmethod
    def portfolio_summary_report(db: Session, period_date: date) -> Dict:
        """Reporte resumen del portafolio."""
        metrics = CalculationEngine.portfolio_metrics(db, period_date)
        composition = CalculationEngine.composition_by_creditor(db, period_date)
        maturity = CalculationEngine.maturity_profile(db)

        return {
            "report_type": "PORTFOLIO_SUMMARY",
            "period_date": period_date.isoformat(),
            "metrics": metrics,
            "composition_by_creditor": composition,
            "maturity_profile": maturity,
        }

    @staticmethod
    def monthly_trend_report(db: Session, months: int = 12) -> List[Dict]:
        """Reporte de tendencia mensual.

        Uses a single GROUP BY query per period_date and creditor_type
        instead of 5 separate queries per date.
        """
        # Get distinct period dates
        dates_sub = (
            db.query(Balance.period_date)
            .filter(Balance.is_active == True)
            .distinct()
            .order_by(Balance.period_date.desc())
            .limit(months)
            .subquery()
        )

        # Single query: all metrics grouped by period_date and creditor_type
        rows = (
            db.query(
                Balance.period_date,
                Creditor.creditor_type,
                func.coalesce(func.sum(Balance.outstanding_usd), 0).label("total_usd"),
                func.sum(
                    case(
                        (Balance.spread_bps.isnot(None), Balance.outstanding_usd * Balance.spread_bps),
                        else_=0,
                    )
                ).label("spread_weight_sum"),
                func.sum(
                    case(
                        (Balance.spread_bps.isnot(None), Balance.outstanding_usd),
                        else_=0,
                    )
                ).label("spread_weight_denom"),
                func.sum(
                    case(
                        (Balance.residual_term_years.isnot(None), Balance.outstanding_usd * Balance.residual_term_years),
                        else_=0,
                    )
                ).label("term_weight_sum"),
                func.sum(
                    case(
                        (Balance.residual_term_years.isnot(None), Balance.outstanding_usd),
                        else_=0,
                    )
                ).label("term_weight_denom"),
            )
            .join(Disbursement, Balance.disbursement_id == Disbursement.id)
            .join(Contract, Disbursement.contract_id == Contract.id)
            .join(Creditor, Contract.creditor_id == Creditor.id)
            .filter(
                Balance.period_date.in_(db.query(dates_sub.c.period_date)),
                Balance.is_active == True,
                Balance.outstanding_usd > 0,
            )
            .group_by(Balance.period_date, Creditor.creditor_type)
            .order_by(Balance.period_date)
            .all()
        )

        # Aggregate by period_date
        period_data: Dict[date, Dict] = {}
        for row in rows:
            pd = row.period_date
            if pd not in period_data:
                period_data[pd] = {
                    "total_usd": Decimal(0),
                    "ifd_usd": Decimal(0),
                    "market_usd": Decimal(0),
                    "spread_sum": Decimal(0),
                    "spread_den": Decimal(0),
                    "term_sum": Decimal(0),
                    "term_den": Decimal(0),
                }
            d = period_data[pd]
            outstanding = Decimal(str(row.total_usd))
            d["total_usd"] += outstanding
            if row.creditor_type == "IFD":
                d["ifd_usd"] += outstanding
            else:
                d["market_usd"] += outstanding
            d["spread_sum"] += Decimal(str(row.spread_weight_sum or 0))
            d["spread_den"] += Decimal(str(row.spread_weight_denom or 0))
            d["term_sum"] += Decimal(str(row.term_weight_sum or 0))
            d["term_den"] += Decimal(str(row.term_weight_denom or 0))

        trend = []
        for pd in sorted(period_data.keys()):
            d = period_data[pd]
            spread = float(d["spread_sum"] / d["spread_den"]) if d["spread_den"] > 0 else None
            term = float(d["term_sum"] / d["term_den"]) if d["term_den"] > 0 else None
            trend.append({
                "period_date": pd.isoformat(),
                "total_usd": float(d["total_usd"]),
                "ifd_usd": float(d["ifd_usd"]),
                "market_usd": float(d["market_usd"]),
                "spread_bps": spread,
                "term_years": term,
            })
        return trend

    @staticmethod
    def covenant_compliance_report(db: Session, period_date: date) -> List[Dict]:
        """Reporte de cumplimiento de covenants."""
        covenants = db.query(Covenant).filter(Covenant.is_active == True).all()
        results = []

        for cov in covenants:
            # Get latest tracking
            tracking = (
                db.query(CovenantTracking)
                .filter(
                    CovenantTracking.covenant_id == cov.id,
                    CovenantTracking.period_date <= period_date,
                )
                .order_by(CovenantTracking.period_date.desc())
                .first()
            )

            results.append({
                "covenant_id": cov.id,
                "name": cov.name,
                "type": cov.covenant_type,
                "limit_value": float(cov.limit_value),
                "unit": cov.unit,
                "current_value": float(tracking.current_value) if tracking else None,
                "utilization_pct": float(tracking.utilization_pct) if tracking and tracking.utilization_pct else None,
                "status": tracking.status if tracking else "SIN_DATOS",
                "last_evaluated": tracking.period_date.isoformat() if tracking else None,
            })
        return results

    @staticmethod
    def _resolve_period_date(db: Session, period_date: date) -> date:
        """Encuentra el period_date real más cercano (último día de mes <= fecha dada)."""
        # Balances use last day of month as period_date
        actual = (
            db.query(func.max(Balance.period_date))
            .filter(
                Balance.outstanding_usd > 0,
                Balance.period_date <= period_date,
            )
            .scalar()
        )
        if actual:
            return actual
        # Fallback: earliest available
        earliest = (
            db.query(func.min(Balance.period_date))
            .filter(Balance.outstanding_usd > 0)
            .scalar()
        )
        return earliest or period_date

    @staticmethod
    def disbursement_detail_report(db: Session, period_date: date) -> List[Dict]:
        """Reporte detallado de todos los desembolsos."""
        # Resolve to actual period_date (last day of month)
        resolved_date = ReportGenerator._resolve_period_date(db, period_date)

        query = (
            db.query(
                Disbursement.disbursement_code,
                Disbursement.disbursement_name,
                Creditor.creditor_type,
                Creditor.short_name.label("creditor_name"),
                Disbursement.amount_usd,
                Balance.outstanding_usd,
                Balance.spread_bps,
                Balance.residual_term_years,
                Disbursement.maturity_date,
                Disbursement.status,
                Contract.amortization_type,
                Contract.interest_rate_type,
                Disbursement.spread_bps_override,
                Contract.spread_bps.label("contract_spread_bps"),
            )
            .join(Contract, Disbursement.contract_id == Contract.id)
            .join(Creditor, Contract.creditor_id == Creditor.id)
            .outerjoin(
                Balance,
                (Balance.disbursement_id == Disbursement.id) & (Balance.period_date == resolved_date),
            )
            .order_by(Creditor.creditor_type, Creditor.short_name, Disbursement.disbursement_code)
        )
        rows = query.all()

        result = []
        for r in rows:
            # Spread: balance > override > contract (fallback chain)
            spread = r.spread_bps
            if not spread:
                spread = r.spread_bps_override if r.spread_bps_override is not None else r.contract_spread_bps

            result.append({
                "code": r.disbursement_code,
                "name": r.disbursement_name,
                "type": r.creditor_type,
                "creditor": r.creditor_name,
                "original_usd": float(r.amount_usd) if r.amount_usd else 0,
                "outstanding_usd": float(r.outstanding_usd) if r.outstanding_usd else 0,
                "spread_bps": float(spread) if spread else None,
                "term_years": float(r.residual_term_years) if r.residual_term_years else None,
                "maturity_date": r.maturity_date.isoformat() if r.maturity_date else None,
                "status": r.status,
                "amort_type": r.amortization_type,
                "rate_type": r.interest_rate_type,
            })
        return result

    @staticmethod
    def generate_csv(data: List[Dict], columns: Optional[List[str]] = None) -> str:
        """Genera contenido CSV a partir de una lista de diccionarios."""
        if not data:
            return ""
        
        if not columns:
            columns = list(data[0].keys())

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        return output.getvalue()

    @staticmethod
    def export_portfolio_csv(db: Session, period_date: date) -> str:
        """Exporta portafolio completo a CSV."""
        data = ReportGenerator.disbursement_detail_report(db, period_date)
        return ReportGenerator.generate_csv(data)

    @staticmethod
    def export_payments_csv(db: Session, disbursement_id: Optional[int] = None) -> str:
        """Exporta calendario de pagos a CSV."""
        query = db.query(PaymentSchedule)
        if disbursement_id:
            query = query.filter(PaymentSchedule.disbursement_id == disbursement_id)
        payments = query.order_by(PaymentSchedule.payment_date).all()

        data = [
            {
                "payment_date": p.payment_date.isoformat(),
                "disbursement_id": p.disbursement_id,
                "type": p.payment_type,
                "amount_original": float(p.amount_original) if p.amount_original else 0,
                "amount_usd": float(p.amount_usd) if p.amount_usd else 0,
                "status": p.status,
            }
            for p in payments
        ]
        return ReportGenerator.generate_csv(data)

    # ------------------------------------------------------------------
    # XLSX Export
    # ------------------------------------------------------------------

    @staticmethod
    def _get_report_data(db: Session, report_type: str, period_date: Optional[date] = None) -> tuple:
        """Returns (title, headers, rows) for a given report type."""
        if period_date is None:
            period_date = date.today()

        if report_type == "portfolio":
            title = f"Portfolio Summary - {period_date.isoformat()}"
            data = ReportGenerator.disbursement_detail_report(db, period_date)
            headers = [
                "Code", "Name", "Type", "Creditor", "Original USD",
                "Outstanding USD", "Spread (bps)", "Term (years)",
                "Maturity Date", "Status", "Amort Type", "Rate Type",
            ]
            rows = [
                [
                    d["code"], d["name"], d["type"], d["creditor"],
                    d["original_usd"], d["outstanding_usd"],
                    d["spread_bps"], d["term_years"],
                    d["maturity_date"], d["status"],
                    d["amort_type"], d["rate_type"],
                ]
                for d in data
            ]

        elif report_type == "payments":
            title = f"Payment Schedule"
            payments = db.query(PaymentSchedule).order_by(PaymentSchedule.payment_date).all()
            headers = [
                "Payment Date", "Disbursement ID", "Type",
                "Amount Original", "Amount USD", "Status",
            ]
            rows = [
                [
                    p.payment_date.isoformat() if p.payment_date else "",
                    p.disbursement_id,
                    p.payment_type,
                    float(p.amount_original) if p.amount_original else 0,
                    float(p.amount_usd) if p.amount_usd else 0,
                    p.status,
                ]
                for p in payments
            ]

        elif report_type == "maturity":
            title = "Maturity Profile"
            data = CalculationEngine.maturity_profile(db)
            headers = [
                "Year", "Creditor Type", "Creditor Name",
                "Instrument Count", "Total Amount USD", "Avg Spread (bps)",
            ]
            rows = [
                [
                    d["year"], d["creditor_type"], d["creditor_name"],
                    d["instrument_count"], d["total_amount_usd"],
                    d["avg_spread_bps"],
                ]
                for d in data
            ]

        elif report_type == "covenants":
            title = f"Covenant Compliance - {period_date.isoformat()}"
            data = ReportGenerator.covenant_compliance_report(db, period_date)
            headers = [
                "Covenant ID", "Name", "Type", "Limit Value",
                "Unit", "Current Value", "Utilization %", "Status",
            ]
            rows = [
                [
                    d["covenant_id"], d["name"], d["type"],
                    d["limit_value"], d["unit"],
                    d["current_value"], d["utilization_pct"],
                    d["status"],
                ]
                for d in data
            ]

        else:
            title = "Report"
            headers = []
            rows = []

        return title, headers, rows

    @staticmethod
    def generate_xlsx(db: Session, report_type: str, period_date: Optional[date] = None) -> bytes:
        """Generate an XLSX file for a given report type and return as bytes."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers

        title, headers, rows = ReportGenerator._get_report_data(db, report_type, period_date)

        wb = Workbook()
        ws = wb.active
        ws.title = report_type.capitalize()

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Header row
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(rows, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, (int, float)) and value is not None:
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="center")

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            col_letter = ws.cell(row=3, column=col_idx).column_letter
            for row_idx in range(3, len(rows) + 4):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_len = max(max_len, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_pdf(db: Session, report_type: str, period_date: Optional[date] = None) -> bytes:
        """Generate a PDF file for a given report type and return as bytes."""
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        title, headers, rows = ReportGenerator._get_report_data(db, report_type, period_date)

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(letter),
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = styles["Title"]
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))

        # Date line
        date_str = (period_date or date.today()).isoformat()
        elements.append(Paragraph(f"Generated: {date_str}", styles["Normal"]))
        elements.append(Spacer(1, 18))

        if headers and rows:
            # Truncate long cell values for PDF rendering
            def truncate(val, max_len=30):
                s = str(val) if val is not None else ""
                return s[:max_len] if len(s) > max_len else s

            table_data = [headers]
            for row in rows:
                table_data.append([truncate(v) for v in row])

            col_count = len(headers)
            available_width = landscape(letter)[0] - 1.0 * inch
            col_width = available_width / col_count

            table = Table(table_data, colWidths=[col_width] * col_count)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No data available for this report.", styles["Normal"]))

        doc.build(elements)
        output.seek(0)
        return output.getvalue()
